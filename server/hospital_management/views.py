import json
import csv
import io
import re
from decimal import Decimal, InvalidOperation

from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.core.cache import cache
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.decorators import api_view, permission_classes,authentication_classes
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db import transaction
from django.db.models import Sum, Q
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.conf import settings
from rest_framework.exceptions import ValidationError
import logging
import json
import os
import tempfile
import requests

from shared.views import sendGenericMailV2

from .models import (
    Companies,
    Department,
    StaffDepartment,
    StaffJobTitle,
    DoctorType,
    AdministrationType,
    TimeSlot,
    DoctorSchedule,
    OPDTicketBooking,
    EmergencyVisit,
    IPDBooking,
    IPDBilling,
    IPDPayment,
    DoctorFees,
    RateChart,
    InsuaranceProvider,
    Patient,
    Bed,
    Ward,
    Room,
    Floor,
    DutyRoaster,
)
from .faker_seed import (
    DEFAULT_HOSPITAL_FAKER_CONFIG,
    seed_hospital_sidebar_faker_data,
)
from company_account.models import CompanyUser
from .serializers import (
    CompanySerializer,
    DepartmentSerializer,
    StaffDepartmentSerializer,
    StaffJobTitleSerializer,
    DoctorTypeSerializer,
    AdministrationTypeSerializer,
    DoctorSerializer,
    AdministrationSerializer,
    DelistedCompanyUserSerializer,
    TimeSlotSerializer,
    DoctorScheduleSerializer,
    OPDTicketBookingSerializer,
    EmergencyVisitSerializer,
    IPDBookingSerializer,
    IPDBillingSerializer,
    IPDPaymentSerializer,
    DoctorFeesSerializer,
    RateChartSerializer,
    InsuaranceProviderSerializer,
    PatientSerializer,
    BedSerializer,
    WardSerializer,
    RoomSerializer,
    DutyRoasterSerializer,
)

from rest_framework import generics, status


def require_company(request):
    """
    Ensure a valid company reference is provided via ?company_id= or request body.
    """
    body_company_ref = None
    if hasattr(request.data, 'get'):
        body_company_ref = (
            request.data.get('company_id')
            or request.data.get('company')
        )
    elif isinstance(request.data, list) and request.data:
        first_item = request.data[0]
        if isinstance(first_item, dict):
            body_company_ref = (
                first_item.get('company_id')
                or first_item.get('company')
            )

    company_ref = (
        request.query_params.get('company_id')
        or body_company_ref
    )
    if isinstance(company_ref, dict):
        company_ref = company_ref.get('id')
    if company_ref in (None, '', []):
        raise ValidationError({'company_id': 'company_id query param or company field is required.'})
    try:
        company_id = int(company_ref)
    except (TypeError, ValueError):
        raise ValidationError({'company_id': 'company_id must be an integer.'})
    try:
        return Companies.objects.get(id=company_id)
    except Companies.DoesNotExist:
        raise ValidationError({'company_id': 'Companies not found.'})


class HospitalSidebarFakerDataView(APIView):
    permission_classes = [AllowAny]
    MAX_COUNT = 500

    def _parse_int(
        self,
        value,
        field_name,
        default,
        *,
        min_value=0,
        max_value=MAX_COUNT,
    ):
        if value in (None, "", "null", "None"):
            return default
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            raise ValidationError({field_name: "Must be an integer."})
        if parsed < min_value or parsed > max_value:
            raise ValidationError(
                {field_name: f"Must be between {min_value} and {max_value}."}
            )
        return parsed

    def _parse_bool(self, value, field_name, default):
        if value in (None, "", "null", "None"):
            return default
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            normalized = value.strip().lower()
            if normalized in ("true", "1", "yes", "y", "on"):
                return True
            if normalized in ("false", "0", "no", "n", "off"):
                return False
        raise ValidationError({field_name: "Must be a boolean."})

    def post(self, request, *args, **kwargs):
        company = require_company(request)
        if hasattr(request.data, "dict"):
            payload = request.data.dict()
        elif isinstance(request.data, dict):
            payload = request.data
        else:
            payload = {}

        config = {}
        for field_name, default_value in DEFAULT_HOSPITAL_FAKER_CONFIG.items():
            raw_value = payload.get(field_name)
            if isinstance(default_value, bool):
                config[field_name] = self._parse_bool(
                    raw_value,
                    field_name,
                    default_value,
                )
            else:
                config[field_name] = self._parse_int(
                    raw_value,
                    field_name,
                    default_value,
                )

        seed = self._parse_int(
            payload.get("seed"),
            "seed",
            None,
            min_value=0,
            max_value=2147483647,
        )
        mode = (payload.get("mode") or "target").strip().lower()
        if mode not in ("target", "append"):
            raise ValidationError({"mode": "Must be either 'target' or 'append'."})
        created_by = (payload.get("created_by") or "faker").strip()[:100]
        default_password = payload.get("default_password") or "abc123"

        try:
            result = seed_hospital_sidebar_faker_data(
                company,
                created_by=created_by,
                default_password=default_password,
                seed=seed,
                config=config,
                mode=mode,
            )
        except ValidationError:
            raise
        except Exception:
            logging.getLogger(__name__).exception(
                "Hospital sidebar faker generation failed for company_id=%s",
                company.id,
            )
            return Response(
                {"error": "Failed to generate faker data."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response(
            {
                "message": "Hospital faker data generated successfully.",
                "company_id": company.id,
                "config": config,
                "result": result,
                "notes": [
                    "Doctor fees are intentionally excluded from this faker API."
                ],
            },
            status=status.HTTP_201_CREATED,
        )


def ensure_auth_user(company_user, password=None, activate=True):
    """
    Ensure a Django auth user exists and is active for the given CompanyUser.
    If missing, create one with a default password (fallback: "abc123").
    """
    User = get_user_model()
    email = (company_user.email or "").strip()
    username = (company_user.username or "").strip()
    phone_number = (
        getattr(company_user, "mobile", None)
        or getattr(company_user, "phone_number", None)
        or ""
    ).strip()
    query = Q()
    if email:
        query |= Q(email__iexact=email) | Q(username__iexact=email)
    if username:
        query |= Q(username__iexact=username)
    if not query:
        return None

    user = User.objects.filter(query).first()
    created = False
    used_password = None
    if not user:
        if not phone_number:
            raise ValidationError(
                {"phone": "Phone is required to create login credentials."}
            )
        if User.objects.filter(phone_number=phone_number).exists():
            raise ValidationError({"phone": "Phone number already exists."})
        default_password = password or getattr(settings, "DEFAULT_NEW_USER_PASSWORD", None) or "abc123"
        login_id = email or username
        user = User.objects.create_user(
            username=login_id,
            email=email or None,
            phone_number=phone_number,
            password=default_password,
            is_active=activate,
        )
        created = True
        used_password = default_password
    else:
        updates = []
        if activate and not user.is_active:
            user.is_active = True
            updates.append("is_active")
        if password:
            user.set_password(password)
            updates.append("password")
            used_password = password
        if updates:
            user.save(update_fields=updates)
    return user, used_password, created


def resolve_default_password(value=None):
    if value:
        return value
    return getattr(settings, "DEFAULT_NEW_USER_PASSWORD", None) or "abc123"


def parse_dict_payload(value, field_name, default=None):
    if value in (None, '', 'null', 'None'):
        return default or {}
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(value)
    except (TypeError, json.JSONDecodeError):
        raise ValidationError({field_name: 'Invalid JSON payload.'})
    if not isinstance(parsed, dict):
        raise ValidationError({field_name: 'Value must be a JSON object.'})
    return parsed


def parse_list_payload(value, field_name, default=None):
    if value in (None, '', 'null', 'None'):
        return default or []
    if isinstance(value, list):
        return value
    try:
        parsed = json.loads(value)
    except (TypeError, json.JSONDecodeError):
        raise ValidationError({field_name: 'Invalid JSON payload.'})
    if not isinstance(parsed, list):
        raise ValidationError({field_name: 'Value must be a JSON array.'})
    return parsed


def normalize_slot_ids(value):
    """
    Accept a single slot ID or a collection and return a cleaned
    list of positive integers with duplicates removed (preserving order).
    """
    if value in (None, '', 'null', 'None'):
        return []
    if isinstance(value, (list, tuple, set)):
        items = list(value)
    else:
        items = [value]
    normalized = []
    seen = set()
    for item in items:
        try:
            slot_id = int(item)
        except (TypeError, ValueError):
            continue
        if slot_id <= 0 or slot_id in seen:
            continue
        normalized.append(slot_id)
        seen.add(slot_id)
    return normalized


def normalize_day_slot_mapping(mapping):
    """
    Ensure day -> [slot_ids] structure regardless of the incoming payload shape.
    """
    if not mapping:
        return {}
    normalized = {}
    for day, slots in mapping.items():
        cleaned = normalize_slot_ids(slots)
        if cleaned:
            normalized[day] = cleaned
    return normalized


def send_account_credentials_email(email, password, role, company_name=None, full_name=None):
    if not email or not password:
        return False
    display_role = (role or "user").strip().title()
    if (role or "").strip().lower() == "administration":
        display_role = "Administration"
    subject = f"Account Credentials at {company_name or 'Hospital'}  "
    greeting_name = full_name or "there"
    message = (
        f"<p>Hello {greeting_name},</p>"
        f"<p>Your account has been created for {company_name or 'the hospital'}.</p>"
        f"<p><strong>Role:</strong> {display_role}</p>"
        f"<p><strong>Username:</strong> {email}</p>"
        f"<p><strong>Password:</strong> {password}</p>"
        "<p>Please reset your password after your first login.</p>"
    )
    try:
        sendGenericMailV2(email, subject, message)
        return True
    except Exception:
        logger.exception("Failed to send account credentials email.")
        return False


def send_approval_email(
    email,
    role,
    company_name=None,
    full_name=None,
    attachment_path="",
):
    if not email:
        return False, "Missing recipient email."
    display_role = (role or "user").strip().title()
    if (role or "").strip().lower() == "administration":
        display_role = "Administration"
    subject = f"User Approved at {company_name or 'Hospital'}  "
    greeting_name = full_name or "there"
    message = (
        f"<p>Hello {greeting_name},</p>"
        f"<p>Your {display_role} account for {company_name or 'the hospital'} has been approved.</p>"
        "<p>You can now log in and start using the system.</p>"
        "<p>If you did not request this account, please contact the hospital admin.</p>"
    )
    try:
        sendGenericMailV2(email, subject, message, attachment_path)
        return True, None
    except Exception:
        logger.exception("Failed to send approval email.")
        return False, "Failed to send approval email."


def build_doctor_pdf_payload(doctor, company):
    schedules = DoctorSchedule.objects.filter(doctor=doctor, companies=company)
    time_slot_labels = {
        slot.id: f"{slot.start_time.strftime('%I:%M %p')} - {slot.end_time.strftime('%I:%M %p')}"
        for slot in TimeSlot.objects.filter(companies=company)
    }
    merged_slots = {}
    merged_slot_labels = {}
    association_type = doctor.association_type or ""
    medical_council_registration = doctor.medical_council_registration or ""
    for schedule in schedules:
        if not association_type and schedule.association_type:
            association_type = schedule.association_type
        mapping = normalize_day_slot_mapping(schedule.available_day_slots or {})
        for day, slots in mapping.items():
            if day not in merged_slots:
                merged_slots[day] = []
            if day not in merged_slot_labels:
                merged_slot_labels[day] = []
            for slot in slots:
                if slot not in merged_slots[day]:
                    merged_slots[day].append(slot)
                    label = time_slot_labels.get(slot)
                    if label:
                        merged_slot_labels[day].append(label)
    return {
        "fullName": getattr(doctor, "name", "") or getattr(doctor, "full_name", ""),
        "email": doctor.email or "",
        "phone": doctor.mobile or "",
        "whatsappNumber": doctor.whatsapp_number or "",
        "associationType": association_type,
        "medicalCouncilRegistration": medical_council_registration,
        "department": doctor.department_id or "",
        "departmentName": getattr(doctor.department, "name", "") if doctor.department_id else "",
        "doctorType": doctor.doctor_type_id or "",
        "doctorTypeName": getattr(doctor.doctor_type, "name", "") if doctor.doctor_type_id else "",
        "availableDaySlots": merged_slots,
        "availableDaySlotLabels": merged_slot_labels,
    }


def generate_doctor_pdf_attachment(prefill_payload, company_id):
    frontend_origin = os.getenv("FRONTEND_ORIGIN", "http://127.0.0.1:93").rstrip("/")
    url = f"{frontend_origin}/hospital-management/api/doctor-pdf"
    params = {
        "prefill": json.dumps(prefill_payload),
        "company_id": str(company_id),
    }
    response = requests.get(url, params=params, timeout=30)
    if response.status_code != 200:
        detail = response.text.strip() if response.text else ""
        suffix = f": {detail}" if detail else ""
        raise ValueError(f"PDF generation failed: {response.status_code}{suffix}")
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    temp_file.write(response.content)
    temp_file.close()
    return temp_file.name


class CompanyListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        companies = Companies.objects.all().order_by('name')
        serializer = CompanySerializer(companies, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = CompanySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class CompanyDetailView(APIView):
    permission_classes = [AllowAny]

    def get_company(self, pk):
        try:
            return Companies.objects.get(pk=pk)
        except Companies.DoesNotExist:
            return None

    def get(self, request, pk):
        company = self.get_company(pk)
        if not company:
            return Response({'error': 'Companies not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CompanySerializer(company)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        company = self.get_company(pk)
        if not company:
            return Response({'error': 'Companies not found.'}, status=status.HTTP_404_NOT_FOUND)
        company.is_active = False
        company.updated_on = timezone.now()
        if request.user and request.user.is_authenticated:
            company.updated_by = str(request.user)
        company.save(update_fields=['is_active', 'updated_on', 'updated_by'])
        return Response({'message': 'Companies deactivated successfully.'}, status=status.HTTP_200_OK)

    def _update(self, request, pk, partial):
        company = self.get_company(pk)
        if not company:
            return Response({'error': 'Companies not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = CompanySerializer(company, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

# Create your views here.

class DepartmentCreate(APIView):
    permission_classes=[AllowAny]
    # @permission_classes([IsAuthenticated])
    
    def post(self, request, *args, **kwargs):
        company = require_company(request)
        data = request.data
        serializer = DepartmentSerializer(data=data)
        
        if serializer.is_valid():
            try:
                with transaction.atomic():
                    serializer.save()
                company_label = getattr(company, "company_name", None) or getattr(company, "name", None)
                send_account_credentials_email(
                    email=data.get("email"),
                    password=data.get("password") or "abc123",
                    role=data.get("role") or "administration",
                    company_name=company_label,
                    full_name=data.get("full_name") or data.get("name"),
                )
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            except Exception as e:
                if 'username' in str(e).lower():
                    return Response(
                        {'username': ['Username already exists. Please choose a different username.']},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                return Response({'error': [str(e)]}, status=status.HTTP_400_BAD_REQUEST)
        
        # Return serializer errors in a structured format
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
    
    
class ApprovedDepartmentList(APIView):
    permission_classes=[AllowAny]
    
    def get(self, request):
        company = require_company(request)
        try:
            departments_list = Department.objects.filter(
                companies=company,
                delist=False,
            ).order_by('-id')
            serializer = DepartmentSerializer(departments_list, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class DepartmentDetail(APIView):
    permission_classes=[AllowAny]

    def get_department(self, pk, company):
        try:
            return Department.objects.get(pk=pk, companies=company)
        except Department.DoesNotExist:
            return None

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        company = require_company(request)
        department = self.get_department(pk, company)
        if not department:
            return Response({'error': 'Department not found.'}, status=status.HTTP_404_NOT_FOUND)

        department.delist = True
        department.delisted_on = timezone.now()
        if request.user and request.user.is_authenticated:
            department.delisted_by = str(request.user)
        department.save(update_fields=['delist', 'delisted_on', 'delisted_by'])
        return Response({'message': 'Department delisted successfully.'}, status=status.HTTP_200_OK)

    def _update(self, request, pk, partial):
        company = require_company(request)
        department = self.get_department(pk, company)
        if not department:
            return Response({'error': 'Department not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = DepartmentSerializer(department, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class StaffDepartmentCreate(APIView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        require_company(request)
        serializer = StaffDepartmentSerializer(data=request.data)

        if serializer.is_valid():
            try:
                with transaction.atomic():
                    serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({'error': [str(e)]}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class ApprovedStaffDepartmentList(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        try:
            departments_list = StaffDepartment.objects.filter(
                companies=company,
                delist=False,
            ).order_by('-id')
            serializer = StaffDepartmentSerializer(departments_list, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class StaffDepartmentDetail(APIView):
    permission_classes = [AllowAny]

    def get_department(self, pk, company):
        try:
            return StaffDepartment.objects.get(pk=pk, companies=company)
        except StaffDepartment.DoesNotExist:
            return None

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        company = require_company(request)
        department = self.get_department(pk, company)
        if not department:
            return Response({'error': 'Staff Department not found.'}, status=status.HTTP_404_NOT_FOUND)

        department.delist = True
        department.delisted_on = timezone.now()
        if request.user and request.user.is_authenticated:
            department.delisted_by = str(request.user)
        department.save(update_fields=['delist', 'delisted_on', 'delisted_by'])
        return Response({'message': 'Staff Department delisted successfully.'}, status=status.HTTP_200_OK)

    def _update(self, request, pk, partial):
        company = require_company(request)
        department = self.get_department(pk, company)
        if not department:
            return Response({'error': 'Staff Department not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = StaffDepartmentSerializer(department, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class StaffJobTitleCreate(APIView):
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        require_company(request)
        serializer = StaffJobTitleSerializer(data=request.data)

        if serializer.is_valid():
            try:
                with transaction.atomic():
                    serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({'error': [str(e)]}, status=status.HTTP_400_BAD_REQUEST)

        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class ApprovedStaffJobTitleList(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        try:
            job_titles = StaffJobTitle.objects.filter(
                companies=company,
                delist=False,
            ).order_by('-id')
            serializer = StaffJobTitleSerializer(job_titles, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class StaffJobTitleDetail(APIView):
    permission_classes = [AllowAny]

    def get_job_title(self, pk, company):
        try:
            return StaffJobTitle.objects.get(pk=pk, companies=company)
        except StaffJobTitle.DoesNotExist:
            return None

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        company = require_company(request)
        job_title = self.get_job_title(pk, company)
        if not job_title:
            return Response({'error': 'Staff Job Title not found.'}, status=status.HTTP_404_NOT_FOUND)

        job_title.delist = True
        job_title.delisted_on = timezone.now()
        if request.user and request.user.is_authenticated:
            job_title.delisted_by = str(request.user)
        job_title.save(update_fields=['delist', 'delisted_on', 'delisted_by'])
        return Response({'message': 'Staff Job Title delisted successfully.'}, status=status.HTTP_200_OK)

    def _update(self, request, pk, partial):
        company = require_company(request)
        job_title = self.get_job_title(pk, company)
        if not job_title:
            return Response({'error': 'Staff Job Title not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = StaffJobTitleSerializer(job_title, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class DoctorTypeCreate(APIView):
    permission_classes=[AllowAny]
    # @permission_classes([IsAuthenticated])
    
    def post(self, request, *args, **kwargs):
        require_company(request)
        serializer = DoctorTypeSerializer(data=request.data)
        
        if serializer.is_valid():
            try:
                with transaction.atomic():
                    serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            except Exception as e:
                if 'username' in str(e).lower():
                    return Response(
                        {'username': ['Username already exists. Please choose a different username.']},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                return Response({'error': [str(e)]}, status=status.HTTP_400_BAD_REQUEST)
        
        # Return serializer errors in a structured format
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
 
 
        
class ApprovedDoctorTypetList(APIView):
    permission_classes=[AllowAny]
    
    def get(self, request):
        company = require_company(request)
        try:
            doctor_type_list = DoctorType.objects.filter(
                companies=company,
                delist=False,
            ).order_by('-id')
            serializer = DoctorTypeSerializer(doctor_type_list, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class DoctorTypeDetail(APIView):
    permission_classes=[AllowAny]

    def get_doctor_type(self, pk, company):
        try:
            return DoctorType.objects.get(pk=pk, companies=company)
        except DoctorType.DoesNotExist:
            return None

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        company = require_company(request)
        doctor_type = self.get_doctor_type(pk, company)
        if not doctor_type:
            return Response({'error': 'Doctor type not found.'}, status=status.HTTP_404_NOT_FOUND)

        doctor_type.delist = True
        doctor_type.delisted_on = timezone.now()
        if request.user and request.user.is_authenticated:
            doctor_type.delisted_by = str(request.user)
        doctor_type.save(update_fields=['delist', 'delisted_on', 'delisted_by'])
        return Response({'message': 'Doctor type delisted successfully.'}, status=status.HTTP_200_OK)

    def _update(self, request, pk, partial):
        company = require_company(request)
        doctor_type = self.get_doctor_type(pk, company)
        if not doctor_type:
            return Response({'error': 'Doctor type not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = DoctorTypeSerializer(doctor_type, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        
class AdministrationTypeCreate(APIView):
    permission_classes=[AllowAny]
    # @permission_classes([IsAuthenticated])
    
    def post(self, request, *args, **kwargs):
        require_company(request)
        serializer = AdministrationTypeSerializer(data=request.data)
        
        if serializer.is_valid():
            try:
                with transaction.atomic():
                    serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            except Exception as e:
                if 'username' in str(e).lower():
                    return Response(
                        {'username': ['Username already exists. Please choose a different username.']},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                return Response({'error': [str(e)]}, status=status.HTTP_400_BAD_REQUEST)
        
        # Return serializer errors in a structured format
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
    
        
class ApprovedAdministratorTypetList(APIView):
    permission_classes=[AllowAny]
    
    def get(self, request):
        company = require_company(request)
        try:
            administrator_type_list = AdministrationType.objects.filter(companies=company, delist=False).order_by('-id')
            serializer=AdministrationTypeSerializer(administrator_type_list, many=True)        
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class AdministrationTypeDetail(APIView):
    permission_classes=[AllowAny]

    def get_administration_type(self, pk, company):
        try:
            return AdministrationType.objects.get(pk=pk, companies=company)
        except AdministrationType.DoesNotExist:
            return None

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        company = require_company(request)
        administration_type = self.get_administration_type(pk, company)
        if not administration_type:
            return Response({'error': 'Administration type not found.'}, status=status.HTTP_404_NOT_FOUND)

        administration_type.delist = True
        administration_type.delisted_on = timezone.now()
        if request.user and request.user.is_authenticated:
            administration_type.delisted_by = str(request.user)
        administration_type.save(update_fields=['delist', 'delisted_on', 'delisted_by'])
        return Response({'message': 'Administration type delisted successfully.'}, status=status.HTTP_200_OK)

    def _update(self, request, pk, partial):
        company = require_company(request)
        administration_type = self.get_administration_type(pk, company)
        if not administration_type:
            return Response({'error': 'Administration type not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = AdministrationTypeSerializer(administration_type, data=request.data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        
        
class DoctorCreateAndGetView(APIView):
    permission_classes=[AllowAny]
    # @permission_classes([IsAuthenticated])
    
    def post(self, request, *args, **kwargs):
        require_company(request)
        serializer = DoctorSerializer(data=request.data, context={'request': request})
        
        if serializer.is_valid():
            try:
                with transaction.atomic():
                    serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            except Exception as e:
                if 'username' in str(e).lower():
                    return Response(
                        {'username': ['Username already exists. Please choose a different username.']},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                return Response({'error': [str(e)]}, status=status.HTTP_400_BAD_REQUEST)
        
        # Return serializer errors in a structured format
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

class DoctorList(APIView):
    permission_classes=[AllowAny]
    
    
    def get(self, request):
        company = require_company(request)
        try:
            role = request.GET.get('role')
            queryset = CompanyUser.objects.filter(company=company, delist=False)
            if role:
                queryset = queryset.filter(role__iexact=role)
            else:
                queryset = queryset.filter(role__iexact="doctor")
            opd_id = request.GET.get('opd_id') or request.GET.get('department_id')
            if opd_id:
                try:
                    opd_id_int = int(opd_id)
                    queryset = queryset.filter(department_id=opd_id_int)
                except (TypeError, ValueError):
                    return Response(
                        {'error': 'Invalid opd_id. It must be a numeric value.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            administrator_type_list = queryset.order_by('-id')
            serializer = DoctorSerializer(
                administrator_type_list,
                many=True,
                context={'request': request},
            )
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class ApprovedDoctorByOpdList(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        opd_id = request.GET.get('opd_id') or request.GET.get('department_id')
        if not opd_id:
            return Response(
                {'error': 'opd_id is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        try:
            opd_id_int = int(opd_id)
        except (TypeError, ValueError):
            return Response(
                {'error': 'Invalid opd_id. It must be a numeric value.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            queryset = (
                DoctorSchedule.objects
                .select_related('doctor', 'companies', 'doctor__department')
                .filter(
                    companies=company,
                    doctor__department_id=opd_id_int,
                    doctor__is_approve=True,
                    doctor__delist=False,
                    doctor__company=company,
                    is_available=True,
                )
                .order_by('-id')
            )
            serializer = DoctorScheduleSerializer(queryset, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class OPDTicketBookingListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        try:
            bookings = OPDTicketBooking.objects.filter(
                companies=company, delist=False
            ).select_related('department', 'doctor')
            search = request.query_params.get('search')
            if search:
                bookings = bookings.filter(
                    Q(name__icontains=search)
                    | Q(mobile__icontains=search)
                    | Q(ticket_no__icontains=search)
                )
            patient_id = request.query_params.get('patient_id')
            if patient_id:
                try:
                    bookings = bookings.filter(patient_id=int(patient_id))
                except ValueError:
                    pass
            limit = request.query_params.get('limit')
            page = request.query_params.get('page')
            include_count = request.query_params.get('include_count') == '1'
            total_count = None
            bookings = bookings.order_by('-created_on')
            if include_count:
                total_count = bookings.count()
            if limit:
                try:
                    limit_value = int(limit)
                    if limit_value > 0:
                        page_value = 1
                        if page:
                            try:
                                page_value = max(1, int(page))
                            except ValueError:
                                page_value = 1
                        offset = (page_value - 1) * limit_value
                        bookings = bookings[offset:offset + limit_value]
                except ValueError:
                    pass
            serializer = OPDTicketBookingSerializer(bookings, many=True)
            if include_count:
                return Response(
                    {'results': serializer.data, 'count': total_count or 0},
                    status=status.HTTP_200_OK
                )
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        print(request.data)
        company = require_company(request)
        data = request.data.copy()
        data['company'] = company.id
        
        for key in ['department', 'doctor']:
            if data.get(key) in ('', 'null', 'None', ''):
                data[key] = None

        serializer = OPDTicketBookingSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class EmergencyVisitListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        try:
            visits = (
                EmergencyVisit.objects
                .filter(companies=company, delist=False)
                .select_related('patient', 'attending_doctor_name')
                .order_by('-created_on')
            )
            serializer = EmergencyVisitSerializer(visits, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        company = require_company(request)
        data = request.data.copy()
        data['company'] = company.id

        for key in ['attending_doctor_name', 'assigned_datetime', 'outcome_datetime', 'visit_datetime']:
            if data.get(key) in ('', 'null', 'None'):
                data[key] = None

        serializer = EmergencyVisitSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class EmergencyVisitDetailView(APIView):
    permission_classes = [AllowAny]

    def get_object(self, pk, company):
        try:
            return EmergencyVisit.objects.select_related(
                'patient', 'attending_doctor_name'
            ).get(pk=pk, companies=company)
        except EmergencyVisit.DoesNotExist:
            return None

    def get(self, request, pk):
        company = require_company(request)
        visit = self.get_object(pk, company)
        if not visit:
            return Response({'error': 'Emergency visit not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = EmergencyVisitSerializer(visit)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def _update(self, request, pk, partial):
        company = require_company(request)
        visit = self.get_object(pk, company)
        if not visit:
            return Response({'error': 'Emergency visit not found.'}, status=status.HTTP_404_NOT_FOUND)
        data = request.data.copy()
        data['company'] = company.id

        for key in ['attending_doctor_name', 'assigned_datetime', 'outcome_datetime', 'visit_datetime']:
            if data.get(key) in ('', 'null', 'None'):
                data[key] = None

        serializer = EmergencyVisitSerializer(visit, data=data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

class OPDTicketBookingDetailView(APIView):
    permission_classes = [AllowAny]

    def get_object(self, pk, company):
        try:
            return OPDTicketBooking.objects.select_related('department', 'doctor').get(
                pk=pk,
                companies=company,
            )
        except OPDTicketBooking.DoesNotExist:
            return None

    def get(self, request, pk):
        company = require_company(request)
        booking = self.get_object(pk, company)
        if not booking:
            return Response({'error': 'Booking not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = OPDTicketBookingSerializer(booking)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        company = require_company(request)
        booking = self.get_object(pk, company)
        if not booking:
            return Response({'error': 'Booking not found.'}, status=status.HTTP_404_NOT_FOUND)
        booking.delist = True
        booking.delisted_on = timezone.now()
        if request.user and request.user.is_authenticated:
            booking.delisted_by = str(request.user)
        booking.save(update_fields=['delist', 'delisted_on', 'delisted_by'])
        return Response({'message': 'Booking delisted successfully.'}, status=status.HTTP_200_OK)

    def _update(self, request, pk, partial):
        company = require_company(request)
        booking = self.get_object(pk, company)
        if not booking:
            return Response({'error': 'Booking not found.'}, status=status.HTTP_404_NOT_FOUND)

        data = request.data.copy()
        for key in ['department', 'doctor']:
            if key in data and data.get(key) in ('', 'null', 'None', ''):
                data[key] = None

        serializer = OPDTicketBookingSerializer(booking, data=data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class PatientListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        try:
            base_qs = Patient.objects.filter(companies=company, delist=False)
            search = request.query_params.get('search')
            if search:
                base_qs = base_qs.filter(
                    Q(name__icontains=search)
                    | Q(mobile__icontains=search)
                    | Q(patient_id__icontains=search)
                )
            limit = request.query_params.get('limit')
            page = request.query_params.get('page')
            include_count = request.query_params.get('include_count') == '1'
            include_index = request.query_params.get('include_index') == '1'
            total_count = None
            patients = base_qs.order_by('-created_on')
            if include_count:
                total_count = patients.count()
            if limit:
                try:
                    limit_value = int(limit)
                    if limit_value > 0:
                        page_value = 1
                        if page:
                            try:
                                page_value = max(1, int(page))
                            except ValueError:
                                page_value = 1
                        offset = (page_value - 1) * limit_value
                        patients = patients[offset:offset + limit_value]
                except ValueError:
                    pass
            serializer = PatientSerializer(patients, many=True)
            payload = serializer.data
            if include_index:
                ordered_ids = list(
                    Patient.objects.filter(companies=company, delist=False)
                    .order_by('-created_on')
                    .values_list('id', flat=True)
                )
                index_map = {pid: idx + 1 for idx, pid in enumerate(ordered_ids)}
                payload = [
                    {**item, 'serial_no': index_map.get(item.get('id'))}
                    for item in payload
                ]
            if include_count:
                return Response(
                    {'results': payload, 'count': total_count or 0},
                    status=status.HTTP_200_OK
                )
            return Response(payload, status=status.HTTP_200_OK)
        except Exception as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        company = require_company(request)
        data = request.data.copy()
        data['companies'] = company.id
        data.pop('company', None)
        data.pop('company_id', None)
        serializer = PatientSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class PatientDetailView(APIView):
    permission_classes = [AllowAny]

    def get_object(self, pk, company):
        try:
            return Patient.objects.get(pk=pk, companies=company)
        except Patient.DoesNotExist:
            return None

    def get(self, request, pk):
        company = require_company(request)
        patient = self.get_object(pk, company)
        if not patient:
            return Response({'error': 'Patient not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = PatientSerializer(patient)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        company = require_company(request)
        patient = self.get_object(pk, company)
        if not patient:
            return Response({'error': 'Patient not found.'}, status=status.HTTP_404_NOT_FOUND)
        patient.delist = True
        patient.delisted_on = timezone.now()
        if request.user and request.user.is_authenticated:
            patient.delisted_by = str(request.user)
        patient.save(update_fields=['delist', 'delisted_on', 'delisted_by'])
        return Response({'message': 'Patient delisted successfully.'}, status=status.HTTP_200_OK)

    def _update(self, request, pk, partial):
        company = require_company(request)
        patient = self.get_object(pk, company)
        if not patient:
            return Response({'error': 'Patient not found.'}, status=status.HTTP_404_NOT_FOUND)
        data = request.data.copy()
        serializer = PatientSerializer(patient, data=data, partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
    
class AdministrationCreateAndGetView(APIView):
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    # @permission_classes([IsAuthenticated])

    def post(self, request, *args, **kwargs):
        company = require_company(request)
        data = request.data.copy()
        data['company'] = company.id
        if data.get('administration_type') in ('', 'null', 'None'):
            data['administration_type'] = None
        if data.get('staff_department') in ('', 'null', 'None'):
            data['staff_department'] = None
        if data.get('staff_job_title') in ('', 'null', 'None'):
            data['staff_job_title'] = None
        serializer = AdministrationSerializer(
            data=data,
            context={'request': request},
        )

        if serializer.is_valid():
            try:
                with transaction.atomic():
                    serializer.save()
                email = data.get("email")
                password = data.get("password") or "abc123"
                full_name = data.get("full_name") or data.get("name") or "there"
                company_label = getattr(company, "company_name", None) or getattr(company, "name", None)
                subject = f"{company_label or 'Hospital'} Administration Account Credentials"
                message = (
                    f"<p>Hello {full_name},</p>"
                    f"<p>Your administration account has been created for {company_label or 'the hospital'}.</p>"
                    "<p><strong>Role:</strong> Administration</p>"
                    f"<p><strong>Username:</strong> {email}</p>"
                    f"<p><strong>Password:</strong> {password}</p>"
                    "<p>Please reset your password after your first login.</p>"
                )
                if email:
                    sendGenericMailV2(email, subject, message)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            except Exception as e:
                if 'username' in str(e).lower():
                    return Response(
                        {'username': ['Username already exists. Please choose a different username.']},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                return Response({'error': [str(e)]}, status=status.HTTP_400_BAD_REQUEST)

        # Return serializer errors in a structured format
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)




class AdministrationList(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        try:
            role = request.GET.get('role')
            administrator_type_list = CompanyUser.objects.filter(
                company=company,
                delist=False,
            )
            if role:
                administrator_type_list = administrator_type_list.filter(
                    role__iexact=role
                )
            else:
                administrator_type_list = administrator_type_list.filter(
                    role__iexact='administration'
                )
            administrator_type_list = administrator_type_list.order_by('-id')
            serializer = AdministrationSerializer(
                administrator_type_list,
                many=True,
                context={'request': request},
            )
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class StaffList(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        try:
            role = request.GET.get('role')
            administrator_type_list = CompanyUser.objects.filter(
                company=company,
                delist=False,
            )
            if role:
                administrator_type_list = administrator_type_list.filter(
                    role__iexact=role
                )
            else:
                administrator_type_list = administrator_type_list.filter(
                    role__iexact='staff'
                )
            administrator_type_list = administrator_type_list.order_by('-id')
            serializer = AdministrationSerializer(
                administrator_type_list,
                many=True,
                context={'request': request},
            )
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class DelistedCompanyUserList(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        try:
            role = request.GET.get('role')
            queryset = CompanyUser.objects.filter(
                company=company,
                delist=True,
            )
            if role:
                queryset = queryset.filter(role__iexact=role)
            queryset = queryset.order_by('-delisted_on', '-id')
            serializer = DelistedCompanyUserSerializer(
                queryset,
                many=True,
                context={'request': request},
            )
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class DelistedCompanyUserRestore(APIView):
    permission_classes = [AllowAny]

    def post(self, request, pk):
        company = require_company(request)
        try:
            user = CompanyUser.objects.filter(
                pk=pk,
                company=company,
                delist=True,
            ).first()
            if not user:
                return Response(
                    {'error': 'Delisted user not found.'},
                    status=status.HTTP_404_NOT_FOUND,
                )
            user.delist = False
            user.delisted_on = None
            user.delisted_by = ""
            if request.user and request.user.is_authenticated:
                user.updated_by = str(request.user)
            user.updated_on = timezone.now()
            user.save(
                update_fields=[
                    'delist',
                    'delisted_on',
                    'delisted_by',
                    'updated_by',
                    'updated_on',
                ]
            )
            return Response(
                {'message': 'User restored successfully.'},
                status=status.HTTP_200_OK,
            )
        except Exception as e:
            print(e)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class StaffDetail(APIView):
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_staff(self, pk, company):
        try:
            return CompanyUser.objects.get(
                pk=pk,
                company=company,
                role__iexact='staff',
            )
        except CompanyUser.DoesNotExist:
            return None

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        company = require_company(request)
        staff = self.get_staff(pk, company)
        if not staff:
            return Response(
                {'error': 'Staff not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        staff.delist = True
        staff.delisted_on = timezone.now()
        if request.user and request.user.is_authenticated:
            staff.delisted_by = str(request.user)
        staff.save(update_fields=['delist', 'delisted_on', 'delisted_by'])
        return Response(
            {'message': 'Staff delisted successfully.'},
            status=status.HTTP_200_OK
        )

    def _update(self, request, pk, partial):
        company = require_company(request)
        staff = self.get_staff(pk, company)
        if not staff:
            return Response(
                {'error': 'Staff not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = AdministrationSerializer(
            staff,
            data=request.data,
            partial=partial,
            context={'request': request},
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

class StaffApprovalView(APIView):
    permission_classes = [AllowAny]

    def post(self, request, staff_id):
        company = require_company(request)
        try:
            staff = CompanyUser.objects.get(
                id=staff_id,
                company=company,
                role__iexact='staff',
            )
        except CompanyUser.DoesNotExist:
            return Response({'error': 'Staff not found.'}, status=status.HTTP_404_NOT_FOUND)

        requested_status = request.data.get('is_approve')
        if requested_status is None:
            requested_status = True
        elif isinstance(requested_status, str):
            requested_status = requested_status.lower() in ('true', '1', 'yes')

        requested_status = bool(requested_status)

        if staff.is_approve == requested_status:
            if requested_status:
                ensure_auth_user(staff, password=request.data.get("password"))[0]
            serialized = AdministrationSerializer(
                staff,
                context={'request': request},
            )
            return Response(serialized.data, status=status.HTTP_200_OK)

        staff.is_approve = requested_status
        staff.is_active = requested_status
        staff.updated_on = timezone.now()
        staff.save(update_fields=['is_approve', 'is_active', 'updated_on'])

        credentials_password = None
        if requested_status:
            credentials_password = resolve_default_password(
                request.data.get("password"),
            )
            ensure_auth_user(
                staff,
                password=credentials_password,
            )

        serialized = AdministrationSerializer(
            staff,
            context={'request': request},
        )
        response_data = dict(serialized.data)
        if requested_status:
            company_label = getattr(company, "company_name", None) or getattr(company, "name", None)
            email_sent, email_error = send_approval_email(
                email=staff.email,
                role=staff.role,
                company_name=company_label,
                full_name=getattr(staff, "name", None) or getattr(staff, "full_name", None),
            )
            response_data["email_sent"] = email_sent
            if email_error:
                response_data["email_error"] = email_error
            if credentials_password:
                credentials_sent = send_account_credentials_email(
                    email=staff.email,
                    password=credentials_password,
                    role=staff.role,
                    company_name=company_label,
                    full_name=getattr(staff, "name", None) or getattr(staff, "full_name", None),
                )
                response_data["credentials_sent"] = credentials_sent
                if not credentials_sent:
                    response_data["credentials_error"] = "Failed to send account credentials email."
        return Response(response_data, status=status.HTTP_200_OK)


class StaffEnsureLoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request, staff_id):
        company = require_company(request)
        try:
            staff = CompanyUser.objects.get(
                id=staff_id,
                company=company,
                role__iexact='staff',
            )
        except CompanyUser.DoesNotExist:
            return Response({'error': 'Staff not found.'}, status=status.HTTP_404_NOT_FOUND)

        payload_email = (request.data.get("email") or "").strip()
        email = payload_email or (staff.email or "").strip()
        if not email:
            return Response({'error': 'Email is required to create login access.'}, status=status.HTTP_400_BAD_REQUEST)

        if email and email != (staff.email or "").strip():
            staff.email = email
            staff.username = email
            staff.updated_on = timezone.now()
            staff.save(update_fields=['email', 'username', 'updated_on'])

        User = get_user_model()
        query = Q(email__iexact=email) | Q(username__iexact=email)
        existing_user = User.objects.filter(query).first()
        created = False
        if not existing_user:
            ensure_auth_user(staff, password=request.data.get("password"))[0]
            created = True
        else:
            ensure_auth_user(staff, password=request.data.get("password"))[0]

        if not staff.can_access:
            staff.can_access = True
            staff.updated_on = timezone.now()
            staff.save(update_fields=['can_access', 'updated_on'])

        serialized = AdministrationSerializer(
            staff,
            context={'request': request},
        )
        return Response(
            {
                'created': created,
                'user_exists': not created,
                'staff': serialized.data,
            },
            status=status.HTTP_200_OK,
        )

class AdministrationDetail(APIView):
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_administration(self, pk, company):
        try:
            return CompanyUser.objects.get(
                pk=pk,
                company=company,
                role__iexact='administration',
            )
        except CompanyUser.DoesNotExist:
            return None

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        company = require_company(request)
        administration = self.get_administration(pk, company)
        if not administration:
            return Response(
                {'error': 'Administration not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        administration.delist = True
        administration.delisted_on = timezone.now()
        if request.user and request.user.is_authenticated:
            administration.delisted_by = str(request.user)
        administration.save(update_fields=['delist', 'delisted_on', 'delisted_by'])
        return Response(
            {'message': 'Administration delisted successfully.'},
            status=status.HTTP_200_OK
        )

    def _update(self, request, pk, partial):
        company = require_company(request)
        administration = self.get_administration(pk, company)
        if not administration:
            return Response(
                {'error': 'Administration not found.'},
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = AdministrationSerializer(
            administration,
            data=request.data,
            partial=partial,
            context={'request': request},
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class DoctorDetail(APIView):
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_doctor(self, pk, company):
        try:
            return CompanyUser.objects.get(
                pk=pk,
                company=company,
                role__iexact="doctor",
            )
        except CompanyUser.DoesNotExist:
            return None

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def delete(self, request, pk):
        company = require_company(request)
        doctor = self.get_doctor(pk, company)
        if not doctor:
            return Response({'error': 'Doctor not found.'}, status=status.HTTP_404_NOT_FOUND)

        doctor.delist = True
        doctor.delisted_on = timezone.now()
        if request.user and request.user.is_authenticated:
            doctor.delisted_by = str(request.user)
        doctor.save(update_fields=['delist', 'delisted_on', 'delisted_by'])
        return Response(
            {'message': 'Doctor delisted successfully.'},
            status=status.HTTP_200_OK
        )

    def _update(self, request, pk, partial):
        company = require_company(request)
        doctor = self.get_doctor(pk, company)
        if not doctor:
            return Response({'error': 'Doctor not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = DoctorSerializer(
            doctor,
            data=request.data,
            partial=partial,
            context={'request': request},
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

class DoctorApprovalView(APIView):
    permission_classes = [AllowAny]

    def post(self, request, doctor_id):
        company = require_company(request)
        try:
            doctor = CompanyUser.objects.get(
                id=doctor_id,
                company=company,
                role__iexact="doctor",
            )
        except CompanyUser.DoesNotExist:
            return Response({'error': 'Doctor not found.'}, status=status.HTTP_404_NOT_FOUND)

        requested_status = request.data.get('is_approve')
        if requested_status is None:
            requested_status = True
        elif isinstance(requested_status, str):
            requested_status = requested_status.lower() in ('true', '1', 'yes')

        requested_status = bool(requested_status)

        if doctor.is_approve == requested_status:
            if requested_status:
                # Ensure auth user is active even if approval state was already true.
                ensure_auth_user(doctor, password=request.data.get("password"))[0]
            serialized = DoctorSerializer(doctor, context={'request': request})
            return Response(serialized.data, status=status.HTTP_200_OK)

        doctor.is_approve = requested_status
        doctor.is_active = requested_status
        doctor.updated_on = timezone.now()
        doctor.save(update_fields=['is_approve', 'is_active', 'updated_on'])

        # Keep auth user active when approving a hospital user.
        credentials_password = None
        if requested_status:
            credentials_password = resolve_default_password(
                request.data.get("password"),
            )
            ensure_auth_user(
                doctor,
                password=credentials_password,
            )

        serialized = DoctorSerializer(doctor, context={'request': request})
        response_data = dict(serialized.data)
        if requested_status:
            company_label = getattr(company, "company_name", None) or getattr(company, "name", None)
            attachment_path = ""
            try:
                prefill_payload = build_doctor_pdf_payload(doctor, company)
                attachment_path = generate_doctor_pdf_attachment(
                    prefill_payload,
                    company_id=company.id,
                )
            except Exception:
                logger.exception("Failed to generate doctor approval PDF.")
                attachment_path = ""
            try:
                email_sent, email_error = send_approval_email(
                    email=doctor.email,
                    role=doctor.role,
                    company_name=company_label,
                    full_name=getattr(doctor, "name", None) or getattr(doctor, "full_name", None),
                    attachment_path=attachment_path,
                )
            finally:
                if attachment_path:
                    try:
                        os.remove(attachment_path)
                    except OSError:
                        logger.exception("Failed to delete temporary PDF file.")
            response_data["email_sent"] = email_sent
            if email_error:
                response_data["email_error"] = email_error
            if credentials_password:
                credentials_sent = send_account_credentials_email(
                    email=doctor.email,
                    password=credentials_password,
                    role=doctor.role,
                    company_name=company_label,
                    full_name=getattr(doctor, "name", None) or getattr(doctor, "full_name", None),
                )
                response_data["credentials_sent"] = credentials_sent
                if not credentials_sent:
                    response_data["credentials_error"] = "Failed to send account credentials email."
        return Response(response_data, status=status.HTTP_200_OK)


class TimeSlotListCreateView(generics.ListCreateAPIView):
    permission_classes = [AllowAny]
    serializer_class = TimeSlotSerializer

    def get_queryset(self):
        company = require_company(self.request)
        return TimeSlot.objects.filter(companies=company).order_by('start_time')

    def get(self, request, *args, **kwargs):
        """GET -> List all available time slots"""
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            "message": "✅ Time slots fetched successfully!",
            "count": len(serializer.data),
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    def post(self, request, *args, **kwargs):
        """POST -> Create a new time slot"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response({
            "message": "✅ Time slot created successfully!",
            "data": serializer.data
        }, status=status.HTTP_201_CREATED)

    def perform_create(self, serializer):
        company = require_company(self.request)
        serializer.save(companies=company)


class TimeSlotDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [AllowAny]
    serializer_class = TimeSlotSerializer

    def get_queryset(self):
        company = require_company(self.request)
        return TimeSlot.objects.filter(companies=company)
        
class DoctorWithScheduleCreateView(APIView):
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def post(self, request):
        try:
            company = require_company(request)
            profile_image = request.FILES.get('profile_image')

            def clean_optional(value):
                if value in (None, '', 'null', 'None'):
                    return None
                return value

            available_day_slots_raw = parse_dict_payload(
                request.data.get('available_day_slots'),
                'available_day_slots',
                default={},
            )
            available_day_slots = normalize_day_slot_mapping(available_day_slots_raw)
            schedule_rows_raw = parse_list_payload(
                request.data.get('schedule_rows'),
                'schedule_rows',
                default=[],
            )
            schedule_rows = schedule_rows_raw or []
            association_type = clean_optional(request.data.get("association_type"))
            medical_council_registration = clean_optional(
                request.data.get("medical_council_registration")
            )
            if isinstance(medical_council_registration, str):
                medical_council_registration = medical_council_registration.strip()

            full_name = request.data.get("full_name") or request.data.get("name")
            email = (request.data.get("email") or "").strip()
            phone = (request.data.get("phone") or request.data.get("mobile") or "").strip()
            username = (email or "").strip().lower()
            if not username:
                raise ValidationError({'email': 'Email is required to create a doctor user.'})
            if not phone:
                raise ValidationError({'phone': 'Phone is required to create a doctor user.'})
            if not str(medical_council_registration or "").strip():
                raise ValidationError(
                    {
                        'medical_council_registration': (
                            'Medical council registration is required to create a doctor user.'
                        )
                    }
                )

            if CompanyUser.objects.filter(company=company, username=username).exists():
                raise ValidationError({'email': 'A user with this email already exists for this company.'})
            if CompanyUser.objects.filter(company=company, email__iexact=email).exists():
                raise ValidationError({'email': 'A doctor with this email already exists for this company.'})
            if CompanyUser.objects.filter(company=company, mobile=phone).exists():
                raise ValidationError({'phone': 'A doctor with this phone already exists for this company.'})

            password = request.data.get("password", "abc123")

            with transaction.atomic():
                department_id = clean_optional(request.data.get("department"))
                doctor_type_id = clean_optional(request.data.get("doctor_type"))
                department = None
                doctor_type = None
                if department_id:
                    department = Department.objects.filter(
                        pk=department_id,
                        companies=company,
                    ).first()
                    if not department:
                        raise ValidationError({'department': 'Department not found.'})
                if doctor_type_id:
                    doctor_type = DoctorType.objects.filter(
                        pk=doctor_type_id,
                        companies=company,
                    ).first()
                    if not doctor_type:
                        raise ValidationError({'doctor_type': 'Doctor type not found.'})

                company_user = CompanyUser.objects.create(
                    company=company,
                    name=full_name or "",
                    username=username,
                    email=email,
                    mobile=phone or "",
                    role=request.data.get("role") or "doctor",
                    whatsapp_number=request.data.get("whatsapp_number"),
                    department=department,
                    doctor_type=doctor_type,
                    association_type=association_type,
                    medical_council_registration=medical_council_registration,
                    created_by=request.data.get("created_by"),
                    is_superuser=False,
                    is_admin=False,
                    is_active=True,
                    is_staff=True,
                    is_owner=False,
                    is_manager=False,
                    is_assistant=True,
                )

                ensure_auth_user(company_user, password=password)

                if profile_image:
                    company_user.profile_image = profile_image
                    company_user.save(update_fields=['profile_image'])

                schedule_payloads = []
                def append_schedule_entry(day, slot_id, assoc_value, is_available_value=True):
                    if slot_id in (None, ''):
                        return
                    schedule_payloads.append({
                        "company": company.id,
                        "association_type": assoc_value,
                        "available_day_slots": {day: slot_id},
                        "is_available": bool(is_available_value),
                        "doctor": company_user.id,
                    })

                def parse_bool(value, default=True):
                    if value in (None, ''):
                        return default
                    if isinstance(value, str):
                        return value.lower() not in ('false', '0', 'no', 'off')
                    return bool(value)

                if schedule_rows:
                    for row in schedule_rows:
                        day = row.get('day')
                        slots = (
                            row.get('time_slots')
                            if row.get('time_slots') not in (None, '', [])
                            else row.get('time_slot') or row.get('slot_id')
                        )
                        if not day or slots in (None, '', []):
                            raise ValidationError({'schedule_rows': 'Each entry requires a day and at least one time slot.'})

                        cleaned_slots = normalize_slot_ids(slots)
                        if not cleaned_slots:
                            raise ValidationError({'schedule_rows': f'Invalid time_slot value for day {day}.'})

                        assoc_value = row.get('association_type', association_type)
                        is_available_value = parse_bool(row.get('is_available', True))

                        for slot_id in cleaned_slots:
                            append_schedule_entry(day, slot_id, assoc_value, is_available_value)
                else:
                    if available_day_slots:
                        for day, slots in available_day_slots.items():
                            cleaned_slots = normalize_slot_ids(slots)
                            if not cleaned_slots:
                                continue
                            for slot_id in cleaned_slots:
                                append_schedule_entry(day, slot_id, association_type, True)
                    else:
                        schedule_payloads.append({
                            "company": company.id,
                            "association_type": association_type,
                            "available_day_slots": {},
                            "is_available": True,
                            "doctor": company_user.id,
                        })

                created_schedules = []
                for payload in schedule_payloads:
                    schedule_serializer = DoctorScheduleSerializer(data=payload)
                    if not schedule_serializer.is_valid():
                        print("Schedule creation error:", schedule_serializer.errors)
                        raise ValidationError({'errors': schedule_serializer.errors})
                    schedule_serializer.save()
                    created_schedules.append(schedule_serializer.data)

                company_label = getattr(company, "company_name", None) or getattr(company, "name", None)
                send_account_credentials_email(
                    email=email,
                    password=password,
                    role=request.data.get("role") or "doctor",
                    company_name=company_label,
                    full_name=full_name,
                )

                return Response({
                    "message": "✅ Doctor and Schedule created successfully!",
                    "doctor": {
                        "id": company_user.id,
                        "full_name": company_user.name,
                        "email": company_user.email,
                        "phone": company_user.mobile,
                        "medical_council_registration": (
                            company_user.medical_council_registration or ""
                        ),
                    },
                    "schedules": created_schedules,
                }, status=status.HTTP_201_CREATED)
        except ValidationError:
            logger.exception("DoctorWithScheduleCreateView validation failed")
            raise
        except Exception:
            logger.exception("DoctorWithScheduleCreateView unexpected error")
            return Response(
                {"error": "Failed to create doctor with schedule."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )







class DoctorScheduleListView(generics.ListCreateAPIView):
    """
    ✅ GET: List all doctor schedules or filter by doctor_id / day.
    Always provide ?company_id=<id> (or company_id in the request body)
    to scope the response to a single company.
    Examples:
      - /api/doctor-schedules/                    → all schedules
      - /api/doctor-schedules/?doctor_id=5        → schedules for doctor 5
      - /api/doctor-schedules/?day=Monday         → schedules available on Monday
      - /api/doctor-schedules/?doctor_id=5&day=Monday → specific doctor + day
    """
    permission_classes = [AllowAny]
    serializer_class = DoctorScheduleSerializer

    def get_queryset(self):
        company = require_company(self.request)
        queryset = DoctorSchedule.objects.select_related("doctor", "companies").filter(
            companies=company,
            delist=False,
        )

        doctor_id = self.request.query_params.get("doctor_id")
        day = self.request.query_params.get("day")

        # ✅ Filter by doctor
        if doctor_id:
            queryset = queryset.filter(doctor_id=doctor_id)

        # ✅ Filter by day key inside JSON
        if day:
            # Postgres JSONField supports __has_key lookup
            queryset = queryset.filter(available_day_slots__has_key=day)

        return queryset.order_by("doctor__name")

    def list(self, request, *args, **kwargs):
        queryset = list(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)

        # ✅ Enrich each schedule with readable time slots
        enriched_data = []
        serialized_rows = serializer.data
        for item in serialized_rows:
            item["readable_slots"] = item.get("readable_schedule") or {}
            enriched_data.append(item)

        return Response(
            {
                "message": "✅ Doctor schedules fetched successfully!",
                "count": len(enriched_data),
                "data": enriched_data,
            },
            status=status.HTTP_200_OK,
        )

    def post(self, request, *args, **kwargs):
        company = require_company(request)
        data = request.data.copy()
        data["company"] = company.id
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save(companies=company)
        return Response(
            {
                "message": "✅ Doctor schedule created successfully!",
                "data": serializer.data,
            },
            status=status.HTTP_201_CREATED,
        )


class DoctorScheduleDetailView(generics.RetrieveUpdateAPIView):
    permission_classes = [AllowAny]
    serializer_class = DoctorScheduleSerializer

    def get_queryset(self):
        company = require_company(self.request)
        return DoctorSchedule.objects.filter(companies=company)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', True)
        instance = self.get_object()
        serializer = self.get_serializer(
            instance,
            data=request.data,
            partial=partial,
        )
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data, status=status.HTTP_200_OK)

        
        
class DutyRoasterListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        try:
            roasters = DutyRoaster.objects.filter(
                companies=company, delist=False
            ).select_related(
                'staff',
                'department',
                'doctor_type',
                'administration_type',
                'staff_department',
                'staff_job_title',
            )

            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            if start_date and end_date:
                roasters = roasters.filter(duty_date__range=[start_date, end_date])
            elif start_date:
                roasters = roasters.filter(duty_date__gte=start_date)
            elif end_date:
                roasters = roasters.filter(duty_date__lte=end_date)

            staff_id = request.query_params.get('staff_id')
            if staff_id:
                roasters = roasters.filter(staff_id=staff_id)

            serializer = DutyRoasterSerializer(roasters.order_by('-duty_date'), many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            print(e)
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        company = require_company(request)
        data = request.data.copy()
        data['companies'] = company.id

        for key in [
            'department',
            'doctor_type',
            'administration_type',
            'staff_department',
            'staff_job_title',
        ]:
            if data.get(key) in ('', 'null', 'None'):
                data[key] = None
        if data.get('time_slot') in ('', 'null', 'None'):
            data['time_slot'] = None
        if data.get('on_call_time') in ('', 'null', 'None'):
            data['on_call_time'] = None
        if data.get('notes') in ('', 'null', 'None'):
            data['notes'] = None

        serializer = DutyRoasterSerializer(data=data, context={'company': company})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class DutyRoasterDetailView(APIView):
    permission_classes = [AllowAny]

    def get_object(self, pk, company):
        try:
            return DutyRoaster.objects.select_related(
                'staff',
                'department',
                'doctor_type',
                'administration_type',
                'staff_department',
                'staff_job_title',
            ).get(pk=pk, companies=company, delist=False)
        except DutyRoaster.DoesNotExist:
            return None

    def get(self, request, pk):
        company = require_company(request)
        roaster = self.get_object(pk, company)
        if not roaster:
            return Response({'error': 'Duty roaster not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = DutyRoasterSerializer(roaster)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, pk):
        company = require_company(request)
        roaster = self.get_object(pk, company)
        if not roaster:
            return Response({'error': 'Duty roaster not found.'}, status=status.HTTP_404_NOT_FOUND)
        data = request.data.copy()
        data['companies'] = company.id
        data.pop('company', None)
        data.pop('company_id', None)

        for key in [
            'department',
            'doctor_type',
            'administration_type',
            'staff_department',
            'staff_job_title',
        ]:
            if data.get(key) in ('', 'null', 'None'):
                data[key] = None
        if data.get('time_slot') in ('', 'null', 'None'):
            data['time_slot'] = None
        if data.get('on_call_time') in ('', 'null', 'None'):
            data['on_call_time'] = None
        if data.get('notes') in ('', 'null', 'None'):
            data['notes'] = None

        serializer = DutyRoasterSerializer(
            roaster,
            data=data,
            partial=True,
            context={'company': company},
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class AdministrationApprovalView(APIView):
    permission_classes = [AllowAny]

    def post(self, request, administration_id):
        company = require_company(request)
        try:
            administration = CompanyUser.objects.get(
                id=administration_id,
                company=company,
                role__iexact='administration',
            )
        except CompanyUser.DoesNotExist:
            return Response({'error': 'Administration not found.'}, status=status.HTTP_404_NOT_FOUND)

        requested_status = request.data.get('is_approve')
        if requested_status is None:
            requested_status = True
        elif isinstance(requested_status, str):
            requested_status = requested_status.lower() in ('true', '1', 'yes')

        requested_status = bool(requested_status)

        if administration.is_approve == requested_status:
            if requested_status:
                # Ensure auth user is active even if approval state was already true.
                ensure_auth_user(administration, password=request.data.get("password"))[0]
            serialized = AdministrationSerializer(
                administration,
                context={'request': request},
            )
            return Response(serialized.data, status=status.HTTP_200_OK)

        administration.is_approve = requested_status
        administration.is_active = requested_status
        administration.updated_on = timezone.now()
        administration.save(update_fields=['is_approve', 'is_active', 'updated_on'])

        # Keep auth user active when approving a hospital user.
        credentials_password = None
        if requested_status:
            credentials_password = resolve_default_password(
                request.data.get("password"),
            )
            ensure_auth_user(
                administration,
                password=credentials_password,
            )

        serialized = AdministrationSerializer(
            administration,
            context={'request': request},
        )
        response_data = dict(serialized.data)
        if requested_status:
            company_label = getattr(company, "company_name", None) or getattr(company, "name", None)
            email_sent, email_error = send_approval_email(
                email=administration.email,
                role=administration.role,
                company_name=company_label,
                full_name=getattr(administration, "name", None) or getattr(administration, "full_name", None),
            )
            response_data["email_sent"] = email_sent
            if email_error:
                response_data["email_error"] = email_error
            if credentials_password:
                credentials_sent = send_account_credentials_email(
                    email=administration.email,
                    password=credentials_password,
                    role=administration.role,
                    company_name=company_label,
                    full_name=getattr(administration, "name", None) or getattr(administration, "full_name", None),
                )
                response_data["credentials_sent"] = credentials_sent
                if not credentials_sent:
                    response_data["credentials_error"] = "Failed to send account credentials email."
        return Response(response_data, status=status.HTTP_200_OK)


class WardListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        wards = Ward.objects.filter(companies=company, delist=False).order_by('name')
        serializer = WardSerializer(wards, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        company = require_company(request)
        data = request.data.copy()
        data['companies'] = company.id
        serializer = WardSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class WardDetailView(APIView):
    permission_classes = [AllowAny]

    def get_object(self, company, pk):
        try:
            return Ward.objects.get(companies=company, pk=pk, delist=False)
        except Ward.DoesNotExist:
            raise ValidationError({'detail': 'Ward not found.'})

    def patch(self, request, pk):
        company = require_company(request)
        ward = self.get_object(company, pk)
        data = request.data.copy()
        data['companies'] = company.id
        serializer = WardSerializer(ward, data=data, partial=True)
        if serializer.is_valid():
            updated = serializer.save()
            output = WardSerializer(updated).data
            return Response(output, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class RoomListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        rooms_qs = (
            Room.objects
            .filter(companies=company, delist=False)
            .select_related('ward', 'floor')
        )
        ward_id = request.query_params.get('ward_id')
        if ward_id:
            rooms_qs = rooms_qs.filter(ward_id=ward_id)
        serializer = RoomSerializer(rooms_qs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        company = require_company(request)
        data = request.data.copy()
        data['companies'] = company.id
        serializer = RoomSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class RoomDetailView(APIView):
    permission_classes = [AllowAny]

    def get_object(self, company, pk):
        try:
            return Room.objects.select_related('ward', 'floor').get(
                companies=company, pk=pk, delist=False
            )
        except Room.DoesNotExist:
            raise ValidationError({'detail': 'Room not found.'})

    def patch(self, request, pk):
        company = require_company(request)
        room = self.get_object(company, pk)
        data = request.data.copy()
        data['companies'] = company.id
        serializer = RoomSerializer(room, data=data, partial=True)
        if serializer.is_valid():
            updated = serializer.save()
            output = RoomSerializer(updated).data
            return Response(output, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class BedListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        try:
            beds = (
                Bed.objects
                .filter(companies=company, delist=False)
                .select_related('room', 'room__ward')
            )
            room_id = request.query_params.get('room_id')
            if room_id:
                beds = beds.filter(room_id=room_id)
            available_only = request.query_params.get('available_only')
            if available_only and available_only.lower() in ('1', 'true', 'yes'):
                beds = beds.filter(is_available=True)
            serializer = BedSerializer(beds, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        company = require_company(request)
        data = request.data.copy()
        data['companies'] = company.id
        if 'is_available' not in data:
            data['is_available'] = True
        serializer = BedSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class BedDetailView(APIView):
    permission_classes = [AllowAny]

    def get_object(self, company, pk):
        try:
            return Bed.objects.select_related('room', 'room__ward').get(
                companies=company, pk=pk, delist=False
            )
        except Bed.DoesNotExist:
            raise ValidationError({'detail': 'Bed not found.'})

    def patch(self, request, pk):
        company = require_company(request)
        bed = self.get_object(company, pk)
        data = request.data.copy()
        data['companies'] = company.id
        serializer = BedSerializer(bed, data=data, partial=True)
        if serializer.is_valid():
            updated = serializer.save()
            output = BedSerializer(updated).data
            return Response(output, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class IPDBookingListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        bookings = (
            IPDBooking.objects
            .filter(companies=company, delist=False)
            .select_related(
                'patient',
                'department',
                'insurance_provider',
                'treatment_doctor',
                'ward',
                'room',
                'bed',
            )
            .prefetch_related('consulting_doctors')
        )
        patient_id = request.query_params.get('patient_id')
        if patient_id:
            try:
                bookings = bookings.filter(patient_id=int(patient_id))
            except ValueError:
                pass
        serializer = IPDBookingSerializer(bookings, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        company = require_company(request)
        data = request.data.copy()
        data['company'] = company.id
        serializer = IPDBookingSerializer(data=data, context={'company': company})
        if serializer.is_valid():
            booking = serializer.save()
            output = IPDBookingSerializer(booking).data
            return Response(output, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class IPDBookingDetailView(APIView):
    permission_classes = [AllowAny]

    def get_object(self, company, pk):
        try:
            return IPDBooking.objects.select_related(
                'patient',
                'department',
                'insurance_provider',
                'treatment_doctor',
                'ward',
                'room',
                'bed',
            ).get(companies=company, pk=pk, delist=False)
        except IPDBooking.DoesNotExist:
            raise ValidationError({'detail': 'IPD booking not found.'})

    def get(self, request, pk):
        company = require_company(request)
        booking = self.get_object(company, pk)
        serializer = IPDBookingSerializer(booking)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, pk):
        company = require_company(request)
        booking = self.get_object(company, pk)
        data = request.data.copy()
        data['company'] = company.id
        serializer = IPDBookingSerializer(
            booking,
            data=data,
            partial=True,
            context={'company': company},
        )
        if serializer.is_valid():
            updated = serializer.save()
            output = IPDBookingSerializer(updated).data
            return Response(output, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class IPDBillingListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        billings = IPDBilling.objects.select_related(
            'ipd_booking',
            'ipd_booking__patient',
            'ipd_booking__department',
            'ipd_booking__rate_chart',
            'ipd_booking__insurance_provider',
        ).filter(companies=company)
        ipd_booking_id = request.query_params.get('ipd_booking_id')
        if ipd_booking_id:
            billings = billings.filter(ipd_booking_id=ipd_booking_id)
        serializer = IPDBillingSerializer(
            billings.order_by('-created_on'), many=True
        )
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        company = require_company(request)
        data = request.data.copy()
        data['company'] = company.id
        serializer = IPDBillingSerializer(
            data=data,
            context={'company': company},
        )
        if serializer.is_valid():
            billing = serializer.save()
            output = IPDBillingSerializer(billing).data
            return Response(output, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class IPDBillingDetailView(APIView):
    permission_classes = [AllowAny]

    def get_object(self, company, pk):
        try:
            return IPDBilling.objects.select_related(
                'ipd_booking',
                'ipd_booking__patient',
                'ipd_booking__department',
                'ipd_booking__rate_chart',
                'ipd_booking__insurance_provider',
            ).get(companies=company, pk=pk)
        except IPDBilling.DoesNotExist:
            raise ValidationError({'detail': 'IPD billing not found.'})

    def get(self, request, pk):
        company = require_company(request)
        billing = self.get_object(company, pk)
        serializer = IPDBillingSerializer(billing)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, pk):
        company = require_company(request)
        billing = self.get_object(company, pk)
        data = request.data.copy()
        data['company'] = company.id
        serializer = IPDBillingSerializer(
            billing,
            data=data,
            partial=True,
            context={'company': company},
        )
        if serializer.is_valid():
            updated = serializer.save()
            output = IPDBillingSerializer(updated).data
            return Response(output, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class IPDPaymentListCreateView(APIView):
    permission_classes = [AllowAny]

    def _sync_billing_status(self, billing):
        if not billing:
            return
        totals = IPDPayment.objects.filter(
            ipd_billing=billing, is_cancelled=False
        ).aggregate(total=Sum('amount'))
        total_paid = totals.get('total') or 0
        due_amount = billing.total_amount - total_paid
        next_status = 'PAID' if due_amount <= 0 else 'PENDING'
        if billing.status != next_status:
            billing.status = next_status
            billing.save(update_fields=['status'])

    def get(self, request):
        company = require_company(request)
        payments = IPDPayment.objects.select_related(
            'ipd_booking',
            'ipd_booking__patient',
            'ipd_billing',
        ).filter(companies=company)
        ipd_booking_id = request.query_params.get('ipd_booking_id')
        if ipd_booking_id:
            payments = payments.filter(ipd_booking_id=ipd_booking_id)
        serializer = IPDPaymentSerializer(
            payments.order_by('-created_on'), many=True
        )
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        company = require_company(request)
        data = request.data.copy()
        data['company'] = company.id
        serializer = IPDPaymentSerializer(
            data=data,
            context={'company': company},
        )
        if serializer.is_valid():
            payment = serializer.save()
            self._sync_billing_status(payment.ipd_billing)
            output = IPDPaymentSerializer(payment).data
            return Response(output, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class IPDPaymentDetailView(APIView):
    permission_classes = [AllowAny]

    def _sync_billing_status(self, billing):
        if not billing:
            return
        totals = IPDPayment.objects.filter(
            ipd_billing=billing, is_cancelled=False
        ).aggregate(total=Sum('amount'))
        total_paid = totals.get('total') or 0
        due_amount = billing.total_amount - total_paid
        next_status = 'PAID' if due_amount <= 0 else 'PENDING'
        if billing.status != next_status:
            billing.status = next_status
            billing.save(update_fields=['status'])

    def get_object(self, company, pk):
        try:
            return IPDPayment.objects.select_related(
                'ipd_booking',
                'ipd_booking__patient',
                'ipd_billing',
            ).get(companies=company, pk=pk)
        except IPDPayment.DoesNotExist:
            raise ValidationError({'detail': 'IPD payment not found.'})

    def get(self, request, pk):
        company = require_company(request)
        payment = self.get_object(company, pk)
        serializer = IPDPaymentSerializer(payment)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, pk):
        company = require_company(request)
        payment = self.get_object(company, pk)
        previous_billing = payment.ipd_billing
        data = request.data.copy()
        data['company'] = company.id
        serializer = IPDPaymentSerializer(
            payment,
            data=data,
            partial=True,
            context={'company': company},
        )
        if serializer.is_valid():
            updated = serializer.save()
            self._sync_billing_status(updated.ipd_billing)
            if previous_billing and previous_billing != updated.ipd_billing:
                self._sync_billing_status(previous_billing)
            output = IPDPaymentSerializer(updated).data
            return Response(output, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class DoctorFeesListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        fees_qs = DoctorFees.objects.select_related(
            'doctor',
            'department',
            'doctor_type',
        ).filter(companies=company, delist=False)
        serializer = DoctorFeesSerializer(
            fees_qs.order_by('doctor__name'),
            many=True,
        )
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        company = require_company(request)
        data = request.data.copy()
        data['company'] = company.id
        serializer = DoctorFeesSerializer(
            data=data,
            context={'company': company},
        )
        if serializer.is_valid():
            fees_entry = serializer.save()
            output = DoctorFeesSerializer(fees_entry).data
            return Response(output, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class DoctorFeesDetailView(APIView):
    permission_classes = [AllowAny]

    def get_object(self, company, pk):
        try:
            return DoctorFees.objects.select_related(
                'doctor',
                'department',
                'doctor_type',
            ).get(companies=company, pk=pk, delist=False)
        except DoctorFees.DoesNotExist:
            raise ValidationError({'detail': 'Doctor fee record not found.'})

    def get(self, request, pk):
        company = require_company(request)
        entry = self.get_object(company, pk)
        serializer = DoctorFeesSerializer(entry)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, pk):
        company = require_company(request)
        entry = self.get_object(company, pk)
        data = request.data.copy()
        data['company'] = company.id
        serializer = DoctorFeesSerializer(
            entry,
            data=data,
            partial=True,
            context={'company': company},
        )
        if serializer.is_valid():
            updated = serializer.save()
            output = DoctorFeesSerializer(updated).data
            return Response(output, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        company = require_company(request)
        entry = self.get_object(company, pk)
        delisted_by = (
            request.data.get('delisted_by')
            if isinstance(request.data, dict)
            else None
        )
        if not delisted_by and request.user and request.user.is_authenticated:
            delisted_by = str(request.user)

        entry.delist = True
        entry.delisted_on = timezone.now()
        if delisted_by:
            entry.delisted_by = str(delisted_by)
        entry.save(update_fields=['delist', 'delisted_on', 'delisted_by'])
        return Response(
            {'message': 'Doctor fee record delisted successfully.'},
            status=status.HTTP_200_OK
        )


class RateChartListCreateView(APIView):
    permission_classes = [AllowAny]

    @staticmethod
    def _normalize_header(header):
        return re.sub(r'[^a-z0-9]', '', str(header or '').strip().lower())

    def _resolve_required_columns(self, headers):
        normalized_headers = [self._normalize_header(header) for header in headers]
        name_index = next(
            (
                index
                for index, value in enumerate(normalized_headers)
                if value == 'name' or value.endswith('name')
            ),
            -1,
        )
        amount_index = next(
            (
                index
                for index, value in enumerate(normalized_headers)
                if value in ('fixedamount', 'amount')
            ),
            -1,
        )
        return name_index, amount_index

    @staticmethod
    def _parse_fixed_amount(raw_amount):
        if isinstance(raw_amount, Decimal):
            amount = raw_amount
        elif isinstance(raw_amount, (int, float)):
            amount = Decimal(str(raw_amount))
        else:
            text_value = str(raw_amount or '').strip().replace(',', '')
            cleaned_value = re.sub(r'[^0-9.-]', '', text_value)
            if cleaned_value in ('', '-', '.', '-.'):
                raise ValueError('Fixed Amount must be numeric.')
            try:
                amount = Decimal(cleaned_value)
            except (InvalidOperation, TypeError, ValueError):
                raise ValueError('Fixed Amount must be numeric.')
        if amount <= 0:
            raise ValueError('Fixed Amount must be greater than zero.')
        return str(amount)

    def _extract_bulk_rows_from_tabular_data(self, rows):
        if not rows:
            return [], ['The uploaded file is empty.']

        headers = rows[0]
        name_index, amount_index = self._resolve_required_columns(headers)
        if name_index == -1 or amount_index == -1:
            return [], ['File must include Name and Fixed Amount columns.']

        payload = []
        errors = []

        for row_number, row in enumerate(rows[1:], start=2):
            if not isinstance(row, (list, tuple)):
                row = list(row or [])

            name_value = row[name_index] if name_index < len(row) else ''
            amount_value = row[amount_index] if amount_index < len(row) else ''

            name = str(name_value or '').strip()
            raw_amount = '' if amount_value is None else amount_value

            if not name and str(raw_amount).strip() == '':
                continue
            if not name:
                errors.append(f'Row {row_number}: Name is required.')
                continue

            try:
                fixed_amount = self._parse_fixed_amount(raw_amount)
            except ValueError as exc:
                errors.append(f'Row {row_number}: {exc}')
                continue

            payload.append(
                {
                    'name': name,
                    'fixed_amount': fixed_amount,
                }
            )

        if not payload and not errors:
            errors.append('No valid rows were found in the uploaded file.')

        return payload, errors

    def _extract_bulk_items_from_file(self, uploaded_file):
        file_name = str(getattr(uploaded_file, 'name', '') or '').lower()
        file_content = uploaded_file.read()
        if not file_content:
            return [], ['Uploaded file is empty.']

        if file_name.endswith('.csv'):
            decoded_content = None
            for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
                try:
                    decoded_content = file_content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if decoded_content is None:
                return [], ['Unable to decode CSV file. Please use UTF-8 or Latin-1 encoding.']

            reader = csv.reader(io.StringIO(decoded_content))
            rows = [row for row in reader if any(str(cell or '').strip() for cell in row)]
            return self._extract_bulk_rows_from_tabular_data(rows)

        if file_name.endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
            except Exception:
                return [], ['XLSX upload is not available because openpyxl is not installed on the server.']

            workbook = None
            try:
                workbook = load_workbook(
                    filename=io.BytesIO(file_content),
                    data_only=True,
                    read_only=True,
                )
                sheet = workbook.active
                rows = [
                    list(row)
                    for row in sheet.iter_rows(values_only=True)
                    if any(cell not in (None, '') for cell in row)
                ]
                return self._extract_bulk_rows_from_tabular_data(rows)
            except Exception:
                return [], ['Unable to read XLSX file. Please verify the file format.']
            finally:
                if workbook is not None:
                    workbook.close()

        return [], ['Only .csv and .xlsx files are supported for bulk upload.']

    def get(self, request):
        company = require_company(request)
        rate_qs = RateChart.objects.select_related(
            'department',
            'insuarance_provider',
        ).filter(companies=company)
        serializer = RateChartSerializer(
            rate_qs.order_by('-created_on'),
            many=True,
        )
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        company = require_company(request)
        if isinstance(request.data, list):
            data = {'items': request.data}
        elif hasattr(request.data, 'copy'):
            data = request.data.copy()
        elif isinstance(request.data, dict):
            data = dict(request.data)
        else:
            data = {}

        data['companies'] = company.id

        uploaded_file = (
            request.FILES.get('bulk_file')
            or request.FILES.get('file')
            or request.FILES.get('upload_file')
        )

        bulk_items = data.get('items') or data.get('rate_charts')
        if isinstance(bulk_items, str):
            try:
                bulk_items = json.loads(bulk_items)
            except json.JSONDecodeError:
                return Response(
                    {'errors': {'items': ['Invalid bulk payload format.']}},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if not isinstance(bulk_items, list) and uploaded_file:
            parsed_rows, parse_errors = self._extract_bulk_items_from_file(uploaded_file)
            if parse_errors:
                return Response(
                    {'errors': {'bulk_file': parse_errors}},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            bulk_items = parsed_rows

        if isinstance(bulk_items, list):
            if not bulk_items:
                return Response(
                    {'errors': {'items': ['At least one row is required for bulk upload.']}},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            defaults = {}
            for field_name in (
                'insuarance_provider',
                'department',
                'effective_from',
                'effective_to',
                'status',
                'is_active',
            ):
                field_value = data.get(field_name)
                if field_value not in (None, ''):
                    defaults[field_name] = field_value

            payload = []
            for index, row in enumerate(bulk_items):
                if not isinstance(row, dict):
                    return Response(
                        {'errors': {'items': [f'Row {index + 1} must be an object.']}},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                payload.append(
                    {
                        **defaults,
                        **row,
                        'companies': company.id,
                    }
                )

            serializer = RateChartSerializer(
                data=payload,
                many=True,
                context={'company': company},
            )
            if serializer.is_valid():
                with transaction.atomic():
                    entries = serializer.save()
                output = RateChartSerializer(entries, many=True).data
                return Response(output, status=status.HTTP_201_CREATED)
            return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        serializer = RateChartSerializer(
            data=data,
            context={'company': company},
        )
        if serializer.is_valid():
            entry = serializer.save()
            output = RateChartSerializer(entry).data
            return Response(output, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class RateChartDetailView(APIView):
    permission_classes = [AllowAny]

    def get_object(self, company, pk):
        try:
            return RateChart.objects.select_related(
                'department',
                'insuarance_provider',
            ).get(companies=company, pk=pk)
        except RateChart.DoesNotExist:
            raise ValidationError({'detail': 'Rate chart record not found.'})

    def get(self, request, pk):
        company = require_company(request)
        entry = self.get_object(company, pk)
        serializer = RateChartSerializer(entry)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, pk):
        company = require_company(request)
        entry = self.get_object(company, pk)
        data = request.data.copy()
        data['companies'] = company.id
        serializer = RateChartSerializer(
            entry,
            data=data,
            partial=True,
            context={'company': company},
        )
        if serializer.is_valid():
            updated = serializer.save()
            output = RateChartSerializer(updated).data
            return Response(output, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class InsuaranceProviderListCreateView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        company = require_company(request)
        entries = InsuaranceProvider.objects.filter(
            companies=company,
            delist=False,
        )
        active_only = str(
            request.query_params.get('active_only', '')
        ).strip().lower() in ('1', 'true', 'yes')
        if active_only:
            entries = entries.filter(status__iexact='active')
        entries = entries.order_by('-created_on')
        serializer = InsuaranceProviderSerializer(entries, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        company = require_company(request)
        data = request.data.copy()
        data['companies'] = company.id
        serializer = InsuaranceProviderSerializer(
            data=data,
            context={'company': company},
        )
        if serializer.is_valid():
            entry = serializer.save()
            output = InsuaranceProviderSerializer(entry).data
            return Response(output, status=status.HTTP_201_CREATED)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class InsuaranceProviderDetailView(APIView):
    permission_classes = [AllowAny]

    def get_object(self, company, pk):
        try:
            return InsuaranceProvider.objects.get(
                companies=company,
                pk=pk,
                delist=False,
            )
        except InsuaranceProvider.DoesNotExist:
            raise ValidationError({'detail': 'Insuarance provider not found.'})

    def get(self, request, pk):
        company = require_company(request)
        entry = self.get_object(company, pk)
        serializer = InsuaranceProviderSerializer(entry)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, pk):
        return self._update(request, pk, partial=False)

    def patch(self, request, pk):
        return self._update(request, pk, partial=True)

    def _update(self, request, pk, partial):
        company = require_company(request)
        entry = self.get_object(company, pk)
        data = request.data.copy()
        data['companies'] = company.id
        serializer = InsuaranceProviderSerializer(
            entry,
            data=data,
            partial=partial,
            context={'company': company},
        )
        if serializer.is_valid():
            updated = serializer.save()
            output = InsuaranceProviderSerializer(updated).data
            return Response(output, status=status.HTTP_200_OK)
        return Response({'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        company = require_company(request)
        entry = self.get_object(company, pk)

        delisted_by = (
            request.data.get('delisted_by')
            if hasattr(request.data, 'get')
            else None
        )
        if not delisted_by and request.user and request.user.is_authenticated:
            delisted_by = str(request.user)

        entry.delist = True
        entry.delisted_on = timezone.now()
        if delisted_by:
            entry.delisted_by = str(delisted_by)
        entry.save(update_fields=['delist', 'delisted_on', 'delisted_by'])
        return Response(
            {'message': 'Insuarance provider delisted successfully.'},
            status=status.HTTP_200_OK,
        )


logger = logging.getLogger(__name__)
